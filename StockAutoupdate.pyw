from bs4 import BeautifulSoup
import requests as req
import openpyxl as xl
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
import threading
import time
from win10toast import ToastNotifier
import win10toast


def notification(msg):
        n = ToastNotifier()
        n.show_toast('Stocks', msg, duration=100)


def getPrice(ticker):
    url = f'https://finance.yahoo.com/quote/{ticker}'
    site = req.get(url).text
    soup = BeautifulSoup(site, 'lxml')
    headline = soup.find('div', class_="My(6px) Pos(r) smartphone_Mt(6px)")
    price = headline.div.span.text
    return price


def process_cell(i, sheet):
    ticker = sheet.cell(row=2, column=i).value
    price = getPrice(ticker)
    sheet.cell(row=sheet.max_row, column=i).value = float(price)


def process_sheet(sheet):
    sheet.cell(column=1, row=sheet.max_row + 1).value = today
    threads = []
    for i in range(2, sheet.max_column + 1):
        t = threading.Thread(target=process_cell, args=(i, sheet))
        t.start()
        threads.append(t)
    for thread in threads:
        thread.join()

if __name__ == '__main__':
        
    start = time.time()
    today = datetime.today().strftime('%d-%b')
    file = xl.load_workbook(r'C:\Users\hassanin\OneDrive\Documents\AutoStocks.xlsx')
    with ThreadPoolExecutor() as exe:
        exe.map(process_sheet, file.worksheets)
    file.save(r'C:\Users\hassanin\OneDrive\Documents\AutoStocks.xlsx')
    file.close()
    end = time.time()
    n = end - start
    
    notification(f'Stocks updated\nTime to finish: {n}')