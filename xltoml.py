import openpyxl

file_path = r'path to xl'

file = openpyxl.load_workbook(file_path)
sheets = file.worksheets
for sheet in sheets:
    with open(f'{sheet.title}.html', 'w') as html:
        html.write(f'''
<!DOCTYPE html>
<html>
    <head>
        <title>{sheet.title}</title>
    </head>
    <body>
        <h1>{sheet.title} stocks</h1>
        <hr>
        <table>
        ''')
        for row in range(1, sheet.max_row + 1):
            html.write('<tr>\n')
            for column in range(1, sheet.max_column + 1):
                html.write(f'<td>{sheet.cell(row=row, column=column).value}</td>\n')
            html.write('</tr>\n')
        html.write('</table>\n</body>\n</html>')